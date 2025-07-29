import flet as ft
import time
from datetime import datetime
import json

class ResultsSummaryView:
    def __init__(self, page, summary_data):
        self.page = page
        self.summary_data = summary_data
        self.session_id = summary_data.get("session_id")
        self.completed_tests = summary_data.get("completed_tests", [])
        self.session_data = summary_data.get("session_data", {})
        
        #  ESTADÍSTICAS GENERALES
        self.total_groups = summary_data.get("total_groups", 0)
        self.total_meters = summary_data.get("total_meters", 0)
        self.total_passed = summary_data.get("total_passed", 0)
        self.total_failed = self.total_meters - self.total_passed
        self.success_rate = (self.total_passed / self.total_meters * 100) if self.total_meters > 0 else 0
        
        print(f"[RESULTS_SUMMARY]  Inicializando vista de resumen")
        print(f"[RESULTS_SUMMARY]  Sesión ID: {self.session_id}")
        print(f"[RESULTS_SUMMARY]  Total grupos: {self.total_groups}")
        print(f"[RESULTS_SUMMARY]  Total medidores: {self.total_meters}")
        print(f"[RESULTS_SUMMARY]  Tasa de éxito: {self.success_rate:.1f}%")

    def _build_general_summary(self):
        """ CONSTRUYE RESUMEN GENERAL CON ESTADÍSTICAS CLAVE"""
        return ft.Container(
            content=ft.Row([
                self._create_stat_card("Grupos de Pruebas", str(self.total_groups), ft.Colors.BLUE_600),
                self._create_stat_card("Total Medidores", str(self.total_meters), ft.Colors.PURPLE_600),
                self._create_stat_card("Medidores Aprobados", str(self.total_passed), ft.Colors.GREEN_600),
                self._create_stat_card("Tasa de Éxito", f"{self.success_rate:.1f}%", 
                                     ft.Colors.GREEN_600 if self.success_rate >= 80 else 
                                     ft.Colors.ORANGE_600 if self.success_rate >= 50 else ft.Colors.RED_600),
            ], alignment=ft.MainAxisAlignment.SPACE_AROUND, spacing=20),
            padding=ft.padding.only(bottom=30),
        )

    def build(self):
        """ CONSTRUYE LA VISTA COMPLETA DE RESUMEN DE RESULTADOS"""
        try:
            #  CREAR TABS PARA ORGANIZAR LA INFORMACIÓN
            tabs = ft.Tabs(
                selected_index=0,
                animation_duration=300,
                tabs=[
                    ft.Tab(
                        text=" Resumen General",
                        content=ft.Container(
                            content=ft.Column([
                                self._build_header(),
                                self._build_general_summary(),
                                self._build_summary_table(),
                            ], spacing=20, scroll=ft.ScrollMode.AUTO),  #  MOVER SCROLL AQUÍ
                            padding=20,
                            height=600,
                        )
                    ),
                    ft.Tab(
                        text=" Detalle de Grupos",
                        content=ft.Container(
                            content=ft.Column([
                                self._build_groups_detail(),
                            ], scroll=ft.ScrollMode.AUTO),  #  MOVER SCROLL AQUÍ
                            height=600,
                        )
                    ),
                    ft.Tab(
                        text="📏 Informe por Medidor", 
                        content=ft.Container(
                            content=ft.Column([
                                self._build_meter_report(),
                            ], scroll=ft.ScrollMode.AUTO),  #  MOVER SCROLL AQUÍ
                            height=600,
                        )
                    ),
                    ft.Tab(
                        text="📈 Estadísticas",
                        content=ft.Container(
                            content=ft.Column([
                                self._build_statistics(),
                            ], scroll=ft.ScrollMode.AUTO),  #  MOVER SCROLL AQUÍ
                            height=600,
                        )
                    ),
                    ft.Tab(
                        text="📁 Exportar",
                        content=ft.Container(
                            content=ft.Column([
                                self._build_export_options(),
                            ], scroll=ft.ScrollMode.AUTO),  #  MOVER SCROLL AQUÍ
                            height=600,
                        )
                    ),
                ],
                height=700,
            )

            #  LAYOUT PRINCIPAL CON ALTURA CONTROLADA
            main_layout = ft.Container(
                content=ft.Column([
                    tabs,
                    self._build_navigation_buttons(),
                ], spacing=0),
                height=800,
                width=None,
            )

            return main_layout
            
        except Exception as e:
            print(f"[RESULTS_SUMMARY]  Error construyendo vista: {e}")
            return self._build_error_view(str(e))

    def _build_header(self):
        """ CONSTRUYE EL HEADER CON INFORMACIÓN COMPLETA DE LA SESIÓN"""
        #  OBTENER INFORMACIÓN COMPLETA DEL BATCH
        batch_info = self.summary_data.get("batch_info", {})
        
        client_name = batch_info.get("client", self.session_data.get("client_name", "Cliente Desconocido"))
        technician_name = batch_info.get("technician", self.session_data.get("technician_name", "Técnico Desconocido"))
        
        #  INFORMACIÓN COMPLETA DEL MEDIDOR
        brand = batch_info.get("meter_brand", self.session_data.get("brand", "N/A"))
        model = batch_info.get("meter_model", self.session_data.get("model", "N/A"))
        meter_type = batch_info.get("meter_type", self.session_data.get("type", "N/A"))
        ratio = batch_info.get("ratio", self.session_data.get("ratio", "N/A"))
        nominal_flow = batch_info.get("nominal_flow", self.session_data.get("nominal_flow", "N/A"))
        diameter = batch_info.get("diameter", self.session_data.get("diameter", "N/A"))
        batch_status = batch_info.get("batch_status", self.session_data.get("batch", "nuevo")).upper()
        
        #  FECHA Y HORA
        current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        #  INDICADOR DE ÉXITO
        success_color = ft.Colors.GREEN if self.success_rate >= 80 else ft.Colors.ORANGE if self.success_rate >= 50 else ft.Colors.RED
        success_icon = "" if self.success_rate >= 80 else "" if self.success_rate >= 50 else ""

        header = ft.Container(
            content=ft.Column([
                #  TÍTULO PRINCIPAL
                ft.Row([
                    ft.Text(
                        " Resumen de Resultados de Pruebas",
                        size=28,
                        weight="bold",
                        color=ft.Colors.BLUE_900
                    ),
                    ft.Container(
                        content=ft.Text(
                            f"{success_icon} {self.success_rate:.1f}%",
                            size=20,
                            weight="bold",
                            color="white"
                        ),
                        bgcolor=success_color,
                        padding=ft.padding.symmetric(horizontal=15, vertical=8),
                        border_radius=20,
                    )
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                
                #  INFORMACIÓN COMPLETA DE LA SESIÓN EN TARJETAS
                ft.ResponsiveRow([
                    #  TARJETA 1: CLIENTE Y TÉCNICO
                    ft.Container(
                        content=ft.Column([
                            ft.Text("👥 Participantes", size=14, weight="bold", color=ft.Colors.BLUE_700),
                            ft.Divider(height=5),
                            ft.Row([
                                ft.Text("👤 Cliente:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(client_name, size=12, color=ft.Colors.BLUE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("🔧 Técnico:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(technician_name, size=12, color=ft.Colors.BLUE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("📅 Fecha:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(current_time, size=12, color=ft.Colors.GREY_600),
                            ]),
                        ], spacing=8),
                        col={"xs": 12, "md": 4},
                        padding=15,
                        bgcolor=ft.Colors.BLUE_50,
                        border_radius=10,
                        border=ft.border.all(1, ft.Colors.BLUE_300),
                        margin=5,
                    ),
                    
                    #  TARJETA 2: ESPECIFICACIONES DEL MEDIDOR
                    ft.Container(
                        content=ft.Column([
                            ft.Text("📏 Especificaciones del Medidor", size=14, weight="bold", color=ft.Colors.PURPLE_700),
                            ft.Divider(height=5),
                            ft.Row([
                                ft.Text("🏭 Marca:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(str(brand), size=12, color=ft.Colors.PURPLE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("🔖 Modelo:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(str(model), size=12, color=ft.Colors.PURPLE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("⚙️ Tipo:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(str(meter_type), size=12, color=ft.Colors.PURPLE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text(" Estado:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(batch_status, size=12, color=ft.Colors.PURPLE_700, weight="bold"),
                            ]),
                        ], spacing=8),
                        col={"xs": 12, "md": 4},
                        padding=15,
                        bgcolor=ft.Colors.PURPLE_50,
                        border_radius=10,
                        border=ft.border.all(1, ft.Colors.PURPLE_300),
                        margin=5,
                    ),
                    
                    #  TARJETA 3: CARACTERÍSTICAS TÉCNICAS
                    ft.Container(
                        content=ft.Column([
                            ft.Text("🔧 Características Técnicas", size=14, weight="bold", color=ft.Colors.ORANGE_700),
                            ft.Divider(height=5),
                            ft.Row([
                                ft.Text("⚡ Ratio:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(f"{ratio}:1" if ratio != "N/A" else "N/A", size=12, color=ft.Colors.ORANGE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("🌊 Caudal Nominal:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(f"{nominal_flow} L/h" if nominal_flow != "N/A" else "N/A", size=12, color=ft.Colors.ORANGE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text("📐 Diámetro:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(f"{diameter} mm" if diameter != "N/A" else "N/A", size=12, color=ft.Colors.ORANGE_700, weight="bold"),
                            ]),
                            ft.Row([
                                ft.Text(" Medidores:", size=12, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text(f"{self.total_meters} probados", size=12, color=ft.Colors.GREEN),
                            ]),
                        ], spacing=8),
                        col={"xs": 12, "md": 4},
                        padding=15,
                        bgcolor=ft.Colors.ORANGE_50,
                        border_radius=10,
                        border=ft.border.all(1, ft.Colors.ORANGE_300),
                        margin=5,
                    ),
                ], spacing=10),
                
            ], spacing=20),
            padding=ft.padding.only(bottom=20),
        )

        return header

    def _build_groups_detail(self):
        """ CONSTRUYE DETALLE COMPLETO DE CADA GRUPO DE PRUEBA"""
        groups_containers = []

        for i, test_group in enumerate(self.completed_tests):
            test_name = test_group["test_name"]
            test_type = test_group["test_type"]
            summary = test_group["summary"]
            results = test_group["results"]

            #  HEADER DEL GRUPO CON MÁS INFORMACIÓN
            group_header = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(f"🧪 {test_name}", size=16, weight="bold", color=ft.Colors.PURPLE_700),
                        ft.Text(f" {summary['passed']}/{summary['total']} ({summary['success_rate']:.1f}%)", 
                                size=14, color=ft.Colors.GREEN if summary['success_rate'] >= 80 else ft.Colors.ORANGE),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    
                    #  INFORMACIÓN ADICIONAL DEL GRUPO
                    ft.Row([
                        ft.Text(f"📅 Completado: {test_group.get('completed_at', 'N/A')}", size=12, color=ft.Colors.GREY_600),
                        ft.Text(f"🔢 Repetición: {test_group.get('repetition', 1)}", size=12, color=ft.Colors.GREY_600),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ], spacing=5),
                padding=15,
                bgcolor=ft.Colors.PURPLE_50,
                border_radius=ft.border_radius.only(top_left=10, top_right=10),
                border=ft.border.all(1, ft.Colors.PURPLE_300),
            )

            #  TABLA DE RESULTADOS DEL GRUPO CON TODOS LOS CAMPOS
            result_rows = []
            for j, result in enumerate(results, 1):
                status_color = ft.Colors.GREEN if result["is_passed"] else ft.Colors.RED
                status_text = " PASA" if result["is_passed"] else " NO PASA"

                #  USAR "serial_number" EN LUGAR DE "serial"
                serial_number = result.get("serial_number", result.get("serial", "N/A"))

                result_rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(j), size=12)),
                    ft.DataCell(ft.Text(serial_number, size=12, weight="bold")),
                    ft.DataCell(ft.Text(f"{result.get('initial_reading', 0):.3f}", size=12)),
                    ft.DataCell(ft.Text(f"{result.get('final_reading', 0):.3f}", size=12)),
                    ft.DataCell(ft.Text(f"{result.get('volume_difference', 0):.3f}L", size=12, color=ft.Colors.BLUE_700)),
                    ft.DataCell(ft.Text(f"{result.get('pattern_volume', 0):.3f}L", size=12, color=ft.Colors.TEAL_700, weight="bold")),
                    ft.DataCell(ft.Text(f"{result.get('error_percentage', 0):.2f}%", size=12, 
                                       color=ft.Colors.RED if abs(result.get('error_percentage', 0)) > 2 else ft.Colors.GREEN)),
                    ft.DataCell(ft.Text(status_text, size=12, color=status_color, weight="bold")),
                ]))

            results_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("#", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Serial", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Lectura Inicial", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Lectura Final", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Diferencia", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Vol. Patrón", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Error %", size=12, weight="bold")),
                    ft.DataColumn(ft.Text("Estado", size=12, weight="bold")),
                ],
                rows=result_rows,
                heading_row_color=ft.Colors.GREY_50,
                data_row_min_height=40,
                column_spacing=15,
            )

            #  CONTAINER DEL GRUPO COMPLETO CON ALTURA CONTROLADA
            group_container = ft.Container(
                content=ft.Column([
                    group_header,
                    ft.Container(
                        content=ft.Column([
                            results_table
                        ], scroll=ft.ScrollMode.AUTO),
                        height=min(250, len(result_rows) * 40 + 80),
                        padding=10,
                        border=ft.border.only(
                            left=ft.BorderSide(1, ft.Colors.PURPLE_300),
                            right=ft.BorderSide(1, ft.Colors.PURPLE_300),
                            bottom=ft.BorderSide(1, ft.Colors.PURPLE_300)
                        ),
                        border_radius=ft.border_radius.only(bottom_left=10, bottom_right=10),
                    ),
                ], spacing=0),
                margin=ft.margin.only(bottom=20),
            )

            groups_containers.append(group_container)

        return ft.Column([
            ft.Text(" Detalle Completo de Grupos de Pruebas", size=18, weight="bold", color=ft.Colors.BLUE_700),
            ft.Column(groups_containers, spacing=0, scroll=ft.ScrollMode.AUTO),
        ], spacing=15)

    def _create_stat_card(self, title, value, color):
        """ CREA UNA TARJETA DE ESTADÍSTICA"""
        return ft.Container(
            content=ft.Column([
                ft.Text(title, size=12, color="white", text_align="center", weight="bold"),
                ft.Text(value, size=24, color="white", text_align="center", weight="bold"),
            ], spacing=10, horizontal_alignment="center"),
            width=180,
            height=100,
            bgcolor=color,
            border_radius=15,
            padding=15,
            alignment=ft.alignment.center,
        )

    def _build_summary_table(self):
        """ CONSTRUYE TABLA RESUMEN POR TIPO DE PRUEBA"""
        #  ORGANIZAR DATOS POR TIPO DE PRUEBA
        summary_by_type = {}
        
        for test_group in self.completed_tests:
            test_type = test_group["test_type"]
            summary = test_group["summary"]
            
            if test_type not in summary_by_type:
                summary_by_type[test_type] = {
                    "groups": 0,
                    "total": 0,
                    "passed": 0,
                    "failed": 0,
                }
            
            summary_by_type[test_type]["groups"] += 1
            summary_by_type[test_type]["total"] += summary["total"]
            summary_by_type[test_type]["passed"] += summary["passed"]
            summary_by_type[test_type]["failed"] += summary["failed"]

        #  CREAR FILAS DE LA TABLA
        table_rows = []
        for test_type in ["Q1", "Q2", "Q3", "Q4"]:
            if test_type in summary_by_type:
                data = summary_by_type[test_type]
                success_rate = (data["passed"] / data["total"] * 100) if data["total"] > 0 else 0
                
                table_rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(test_type, weight="bold", color=ft.Colors.PURPLE_700)),
                    ft.DataCell(ft.Text(str(data["groups"]))),
                    ft.DataCell(ft.Text(str(data["total"]))),
                    ft.DataCell(ft.Text(str(data["passed"]), color=ft.Colors.GREEN)),
                    ft.DataCell(ft.Text(str(data["failed"]), color=ft.Colors.RED)),
                    ft.DataCell(ft.Text(f"{success_rate:.1f}%", weight="bold", 
                                       color=ft.Colors.GREEN if success_rate >= 80 else ft.Colors.ORANGE)),
                ]))

        summary_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Tipo", weight="bold")),
                ft.DataColumn(ft.Text("Grupos", weight="bold")),
                ft.DataColumn(ft.Text("Total", weight="bold")),
                ft.DataColumn(ft.Text("Aprobados", weight="bold")),
                ft.DataColumn(ft.Text("Reprobados", weight="bold")),
                ft.DataColumn(ft.Text("Éxito %", weight="bold")),
            ],
            rows=table_rows,
            border=ft.border.all(1, ft.Colors.GREY_300),
            heading_row_color=ft.Colors.BLUE_50,
        )

        return ft.Container(
            content=ft.Column([
                ft.Text(" Resumen por Tipo de Prueba", size=18, weight="bold", color=ft.Colors.BLUE_700),
                ft.Container(
                    content=summary_table,
                    border_radius=10,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    padding=10,
                ),
            ], spacing=10),
        )


    def _build_meter_report(self):
        """ CONSTRUYE INFORME POR MEDIDOR INDIVIDUAL CON TODOS LOS CAMPOS"""
        #  ORGANIZAR DATOS POR MEDIDOR
        meters_data = {}
        
        for test_group in self.completed_tests:
            for result in test_group["results"]:
                #  USAR "serial_number" EN LUGAR DE "serial"
                serial = result.get("serial_number", result.get("serial", "N/A"))
                
                if serial not in meters_data:
                    meters_data[serial] = {
                        "serial": serial,
                        "tests": [],
                        "total_tests": 0,
                        "passed_tests": 0,
                        "failed_tests": 0,
                        "overall_passed": True,
                    }
                
                meters_data[serial]["tests"].append({
                    "test_name": test_group["test_name"],
                    "test_type": test_group["test_type"],
                    "result": result,
                })
                
                meters_data[serial]["total_tests"] += 1
                if result["is_passed"]:
                    meters_data[serial]["passed_tests"] += 1
                else:
                    meters_data[serial]["failed_tests"] += 1
                    meters_data[serial]["overall_passed"] = False

        #  CREAR TARJETAS PARA CADA MEDIDOR
        meter_cards = []
        
        for serial, meter_data in meters_data.items():
            overall_color = ft.Colors.GREEN if meter_data["overall_passed"] else ft.Colors.RED
            overall_icon = "" if meter_data["overall_passed"] else ""
            
            #  TABLA DE PRUEBAS DEL MEDIDOR CON TODOS LOS CAMPOS
            test_rows = []
            for test in meter_data["tests"]:
                result = test["result"]
                status_color = ft.Colors.GREEN if result["is_passed"] else ft.Colors.RED
                status_text = "PASA" if result["is_passed"] else "NO PASA"
                
                test_rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(test["test_name"], size=11)),
                    ft.DataCell(ft.Text(test["test_type"], size=11, weight="bold", color=ft.Colors.PURPLE_700)),
                    ft.DataCell(ft.Text(f"{result.get('initial_reading', 0):.3f}", size=11)),
                    ft.DataCell(ft.Text(f"{result.get('final_reading', 0):.3f}", size=11)),
                    ft.DataCell(ft.Text(f"{result.get('volume_difference', 0):.3f}L", size=11, color=ft.Colors.BLUE_700)),
                    ft.DataCell(ft.Text(f"{result.get('pattern_volume', 0):.3f}L", size=11, color=ft.Colors.TEAL_700, weight="bold")),
                    ft.DataCell(ft.Text(f"{result.get('error_percentage', 0):.2f}%", size=11, 
                                       color=ft.Colors.RED if abs(result.get('error_percentage', 0)) > 2 else ft.Colors.GREEN)),
                    ft.DataCell(ft.Text(status_text, size=11, color=status_color, weight="bold")),
                ]))

            #  TABLA CON TODAS LAS COLUMNAS
            meter_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("Prueba", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("Tipo", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("L. Inicial", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("L. Final", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("Diferencia", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("Vol. Patrón", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("Error %", size=11, weight="bold")),
                    ft.DataColumn(ft.Text("Estado", size=11, weight="bold")),
                ],
                rows=test_rows,
                heading_row_color=ft.Colors.GREY_50,
                data_row_min_height=35,
                column_spacing=8,
            )

            #  TARJETA DEL MEDIDOR CON ALTURA CONTROLADA
            meter_card = ft.Container(
                content=ft.Column([
                    #  HEADER DEL MEDIDOR CON MÁS INFORMACIÓN
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Text(f"📏 Serial: {serial}", size=16, weight="bold", color="white"),
                                ft.Text(f"{overall_icon} {meter_data['passed_tests']}/{meter_data['total_tests']}", 
                                        size=14, color="white", weight="bold"),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            
                            #  ESTADÍSTICAS ADICIONALES DEL MEDIDOR
                            ft.Row([
                                ft.Text(f"Aprobadas: {meter_data['passed_tests']}", size=12, color="white"),
                                ft.Text(f"Reprobadas: {meter_data['failed_tests']}", size=12, color="white"),
                                ft.Text(f"Total: {meter_data['total_tests']}", size=12, color="white"),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ], spacing=5),
                        padding=15,
                        bgcolor=overall_color,
                        border_radius=ft.border_radius.only(top_left=10, top_right=10),
                    ),
                    
                    #  TABLA DE PRUEBAS CON ALTURA FIJA
                    ft.Container(
                        content=ft.Row([meter_table], scroll=ft.ScrollMode.AUTO),
                        height=min(200, len(test_rows) * 35 + 60),  #  ALTURA MÁXIMA CONTROLADA
                        padding=10,
                        border=ft.border.only(
                            left=ft.BorderSide(1, overall_color),
                            right=ft.BorderSide(1, overall_color),
                            bottom=ft.BorderSide(1, overall_color)
                        ),
                        border_radius=ft.border_radius.only(bottom_left=10, bottom_right=10),
                    ),
                ], spacing=0),
                margin=ft.margin.only(bottom=20),
            )

            meter_cards.append(meter_card)

        return ft.Column([
            ft.Text(" Informe Detallado por Medidor Individual", size=18, weight="bold", color=ft.Colors.BLUE_700),
            ft.Text(f"Total de medidores probados: {len(meter_cards)}", size=14, color=ft.Colors.GREY_600),
            
            #  DESCRIPCIÓN DE LAS COLUMNAS
            ft.Container(
                content=ft.Column([
                    ft.Text(" Campos del reporte:", size=14, weight="bold", color=ft.Colors.PURPLE_700),
                    ft.Row([
                        ft.Text("• Prueba: Nombre de la prueba realizada", size=12, color=ft.Colors.GREY_600),
                        ft.Text("• Tipo: Q1, Q2, Q3 o Q4", size=12, color=ft.Colors.GREY_600),
                    ]),
                    ft.Row([
                        ft.Text("• L. Inicial/Final: Lecturas del medidor", size=12, color=ft.Colors.GREY_600),
                        ft.Text("• Diferencia: Volumen medido", size=12, color=ft.Colors.GREY_600),
                    ]),
                    ft.Row([
                        ft.Text("• Vol. Patrón: Volumen de referencia", size=12, color=ft.Colors.GREY_600),
                        ft.Text("• Error %: Porcentaje de error calculado", size=12, color=ft.Colors.GREY_600),
                    ]),
                ], spacing=5),
                padding=15,
                bgcolor=ft.Colors.BLUE_50,
                border_radius=10,
                margin=ft.margin.only(bottom=20),
            ),
            
            #  CONTENEDOR CON LAS TARJETAS SIN EXPAND
            ft.Column(meter_cards, spacing=0),
        ], spacing=15)

    def _build_statistics(self):
        """ CONSTRUYE VISTA DE ESTADÍSTICAS"""
        return ft.Column([
            ft.Text("📈 Estadísticas Detalladas", size=18, weight="bold", color=ft.Colors.BLUE_700),
            ft.Text("🚧 Próximamente: Gráficos y análisis estadísticos", size=16, color=ft.Colors.GREY_600),
            
            #  ESTADÍSTICAS BÁSICAS POR AHORA
            ft.Container(
                content=ft.Column([
                    ft.Text(" Estadísticas Básicas", size=16, weight="bold", color=ft.Colors.PURPLE_700),
                    ft.Row([
                        ft.Text(f"• Total de sesiones probadas: {self.total_groups}"),
                        ft.Text(f"• Promedio de medidores por sesión: {self.total_meters/max(self.total_groups,1):.1f}"),
                    ]),
                    ft.Row([
                        ft.Text(f"• Tasa de éxito general: {self.success_rate:.1f}%"),
                        ft.Text(f"• Medidores reprobados: {self.total_failed}"),
                    ]),
                ], spacing=10),
                padding=20,
                bgcolor=ft.Colors.GREY_50,
                border_radius=10,
            ),
        ], spacing=20)

    def _build_export_options(self):
        """ CONSTRUYE OPCIONES DE EXPORTACIÓN"""
        return ft.Column([
            ft.Text("📁 Opciones de Exportación", size=18, weight="bold", color=ft.Colors.BLUE_700),
            
            #  OPCIONES DE EXPORTACIÓN
            ft.Container(
                content=ft.Column([
                    ft.Text("Seleccione el formato de exportación:", size=14, weight="bold"),
                    
                    ft.Row([
                        ft.ElevatedButton(
                            content=ft.Row([
                                ft.Icon(ft.Icons.TABLE_VIEW, color="white"),
                                ft.Text("Exportar a Excel", color="white"),
                            ], spacing=10),
                            bgcolor=ft.Colors.GREEN_600,
                            on_click=self._export_to_excel,
                            width=200,
                            height=50,
                        ),
                        
                        ft.ElevatedButton(
                            content=ft.Row([
                                ft.Icon(ft.Icons.PICTURE_AS_PDF, color="white"),
                                ft.Text("Exportar a PDF", color="white"),
                            ], spacing=10),
                            bgcolor=ft.Colors.RED_600,
                            on_click=self._export_to_pdf,
                            width=200,
                            height=50,
                        ),
                    ], alignment=ft.MainAxisAlignment.CENTER, spacing=30),
                    
                    ft.Divider(height=20),
                    
                    ft.Text(" El reporte incluirá:", size=14, weight="bold"),
                    ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=16),
                            ft.Text("Datos completos del cliente y técnico", size=12),
                        ], spacing=10),
                        ft.Row([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=16),
                            ft.Text("Especificaciones técnicas del medidor", size=12),
                        ], spacing=10),
                        ft.Row([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=16),
                            ft.Text("Todos los campos de las pruebas realizadas", size=12),
                        ], spacing=10),
                        ft.Row([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=16),
                            ft.Text("Estadísticas y análisis de resultados", size=12),
                        ], spacing=10),
                    ], spacing=5),
                    
                ], spacing=20, horizontal_alignment="center"),
                padding=30,
                bgcolor=ft.Colors.GREY_50,
                border_radius=15,
                alignment=ft.alignment.center,
            ),
            
        ], spacing=20, horizontal_alignment="center")

    def _build_navigation_buttons(self):
        """ CONSTRUYE BOTONES DE NAVEGACIÓN"""
        return ft.Container(
            content=ft.Row([
                ft.ElevatedButton(
                    content=ft.Row([
                        ft.Icon(ft.Icons.ARROW_BACK),
                        ft.Text("Nueva Sesión"),
                    ], spacing=10),
                    bgcolor=ft.Colors.BLUE_600,
                    color="white",
                    on_click=self._start_new_session,
                    width=150,
                ),
                
                ft.ElevatedButton(
                    content=ft.Row([
                        ft.Icon(ft.Icons.HOME),
                        ft.Text("Inicio"),
                    ], spacing=10),
                    bgcolor=ft.Colors.GREY_600,
                    color="white",
                    on_click=self._go_to_home,
                    width=150,
                ),
                
                ft.ElevatedButton(
                    content=ft.Row([
                        ft.Icon(ft.Icons.PRINT),
                        ft.Text("Imprimir"),
                    ], spacing=10),
                    bgcolor=ft.Colors.GREEN_600,
                    color="white",
                    on_click=self._print_report,
                    width=150,
                ),
            ], alignment=ft.MainAxisAlignment.CENTER, spacing=20),
            padding=20,
        )

    def _build_error_view(self, error_message):
        """ CONSTRUYE VISTA DE ERROR"""
        return ft.Container(
            content=ft.Column([
                ft.Icon(ft.Icons.ERROR, size=64, color=ft.Colors.RED),
                ft.Text("Error al cargar el resumen", size=24, weight="bold", color=ft.Colors.RED),
                ft.Text(f"Error: {error_message}", size=14, color=ft.Colors.GREY_600),
                ft.ElevatedButton(
                    "Volver al Inicio",
                    on_click=self._go_to_home,
                    bgcolor=ft.Colors.BLUE_600,
                    color="white",
                ),
            ], horizontal_alignment="center", spacing=20),
            alignment=ft.alignment.center,
            expand=True,
        )

    def _export_to_excel(self, e):
        """ EXPORTA A EXCEL CON SELECCIÓN DE UBICACIÓN"""
        print("[RESULTS_SUMMARY]  Iniciando exportación a Excel...")
        
        def on_file_picker_result(result):
            if result.path:
                try:
                    self._generate_excel_report(result.path)
                    self._show_notification(f" Excel generado exitosamente en: {result.path}")
                except Exception as ex:
                    self._show_notification(f" Error generando Excel: {str(ex)}")
            else:
                self._show_notification(" Exportación cancelada")
        
        #  CREAR FILE PICKER
        file_picker = ft.FilePicker(
            on_result=on_file_picker_result
        )
        
        #  AGREGAR AL OVERLAY DE LA PÁGINA
        self.page.overlay.append(file_picker)
        self.page.update()
        
        #  ABRIR DIÁLOGO PARA GUARDAR
        file_picker.save_file(
            dialog_title="Guardar reporte Excel",
            file_name=f"Reporte_Medidores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            allowed_extensions=["xlsx"]
        )

    def _export_to_pdf(self, e):
        """ EXPORTA A PDF CON SELECCIÓN DE UBICACIÓN"""
        print("[RESULTS_SUMMARY] 📄 Iniciando exportación a PDF...")
        
        def on_file_picker_result(result):
            if result.path:
                try:
                    self._generate_pdf_report(result.path)
                    self._show_notification(f"📄 PDF generado exitosamente en: {result.path}")
                except Exception as ex:
                    self._show_notification(f" Error generando PDF: {str(ex)}")
            else:
                self._show_notification(" Exportación cancelada")
        
        #  CREAR FILE PICKER
        file_picker = ft.FilePicker(
            on_result=on_file_picker_result
        )
        
        #  AGREGAR AL OVERLAY DE LA PÁGINA
        self.page.overlay.append(file_picker)
        self.page.update()
        
        #  ABRIR DIÁLOGO PARA GUARDAR
        file_picker.save_file(
            dialog_title="Guardar reporte PDF",
            file_name=f"Reporte_Medidores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            allowed_extensions=["pdf"]
        )

    def _generate_excel_report(self, file_path):
        """ GENERA REPORTE COMPLETO EN EXCEL"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            #  CREAR WORKBOOK
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Pruebas"
            
            #  CONFIGURAR ESTILOS
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            data_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            #  TÍTULO PRINCIPAL
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "REPORTE DE PRUEBAS DE MEDIDORES"
            ws[f'A{row}'].font = Font(bold=True, size=16)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            #  INFORMACIÓN DE LA SESIÓN
            batch_info = self.summary_data.get("batch_info", {})
            session_info = [
                ("Fecha del Reporte:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                ("ID de Sesión:", str(self.session_id)),
                ("Cliente:", batch_info.get("client", "N/A")),
                ("Técnico:", batch_info.get("technician", "N/A")),
                ("Marca del Medidor:", batch_info.get("meter_brand", "N/A")),
                ("Modelo del Medidor:", batch_info.get("meter_model", "N/A")),
                ("Tipo de Medidor:", batch_info.get("meter_type", "N/A")),
                ("Ratio:", f"{batch_info.get('ratio', 'N/A')}:1"),
                ("Caudal Nominal:", f"{batch_info.get('nominal_flow', 'N/A')} L/h"),
                ("Diámetro:", f"{batch_info.get('diameter', 'N/A')} mm"),
                ("Estado del Lote:", batch_info.get("batch_status", "N/A")),
            ]
            
            for label, value in session_info:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            row += 2
            
            #  ESTADÍSTICAS GENERALES
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "ESTADÍSTICAS GENERALES"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            stats_data = [
                ("Total Grupos de Pruebas:", self.total_groups),
                ("Total Medidores Probados:", self.total_meters),
                ("Medidores Aprobados:", self.total_passed),
                ("Medidores Reprobados:", self.total_failed),
                ("Tasa de Éxito:", f"{self.success_rate:.1f}%"),
            ]
            
            for label, value in stats_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            row += 2
            
            #  DETALLES POR GRUPO DE PRUEBAS
            for test_group in self.completed_tests:
                # Header del grupo
                ws.merge_cells(f'A{row}:H{row}')
                ws[f'A{row}'] = f"GRUPO: {test_group['test_name']}"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                row += 1
                
                # Headers de columnas
                headers = ["#", "Serial", "Lectura Inicial", "Lectura Final", "Diferencia (L)", "Vol. Patrón (L)", "Error (%)", "Estado"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                row += 1
                
                # Datos del grupo
                for idx, result in enumerate(test_group["results"], 1):
                    serial_number = result.get("serial_number", result.get("serial", "N/A"))
                    data_row = [
                        idx,
                        serial_number,
                        f"{result.get('initial_reading', 0):.3f}",
                        f"{result.get('final_reading', 0):.3f}",
                        f"{result.get('volume_difference', 0):.3f}",
                        f"{result.get('pattern_volume', 0):.3f}",
                        f"{result.get('error_percentage', 0):.2f}",
                        "PASA" if result.get("is_passed", False) else "NO PASA"
                    ]
                    
                    for col, value in enumerate(data_row, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = data_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Color según el estado
                        if col == 8:  # Columna de estado
                            if value == "PASA":
                                cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                    
                    row += 1
                
                row += 1  # Espacio entre grupos
            
            #  AJUSTAR ANCHO DE COLUMNAS
            column_widths = [5, 15, 15, 15, 15, 15, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            #  GUARDAR ARCHIVO
            wb.save(file_path)
            print(f"[RESULTS_SUMMARY]  Excel generado en: {file_path}")
            
        except ImportError:
            raise Exception("Se requiere instalar 'pandas' y 'openpyxl' para exportar a Excel:\npip install pandas openpyxl")
        except Exception as e:
            raise Exception(f"Error generando Excel: {str(e)}")

    def _generate_pdf_report(self, file_path):
        """ GENERA REPORTE COMPLETO EN PDF"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            #  CREAR DOCUMENTO
            doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=0.5*inch)
            styles = getSampleStyleSheet()
            story = []
            
            #  TÍTULO
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph("REPORTE DE PRUEBAS DE MEDIDORES", title_style))
            
            #  INFORMACIÓN DE LA SESIÓN
            batch_info = self.summary_data.get("batch_info", {})
            session_data = [
                ["Fecha del Reporte:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["ID de Sesión:", str(self.session_id)],
                ["Cliente:", batch_info.get("client", "N/A")],
                ["Técnico:", batch_info.get("technician", "N/A")],
                ["Marca del Medidor:", batch_info.get("meter_brand", "N/A")],
                ["Modelo del Medidor:", batch_info.get("meter_model", "N/A")],
                ["Tipo de Medidor:", batch_info.get("meter_type", "N/A")],
                ["Ratio:", f"{batch_info.get('ratio', 'N/A')}:1"],
                ["Caudal Nominal:", f"{batch_info.get('nominal_flow', 'N/A')} L/h"],
                ["Diámetro:", f"{batch_info.get('diameter', 'N/A')} mm"],
                ["Estado del Lote:", batch_info.get("batch_status", "N/A")],
            ]
            
            session_table = Table(session_data, colWidths=[2.5*inch, 3*inch])
            session_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(session_table)
            story.append(Spacer(1, 20))
            
            #  ESTADÍSTICAS GENERALES
            story.append(Paragraph("ESTADÍSTICAS GENERALES", styles['Heading2']))
            stats_data = [
                ["Total Grupos de Pruebas:", str(self.total_groups)],
                ["Total Medidores Probados:", str(self.total_meters)],
                ["Medidores Aprobados:", str(self.total_passed)],
                ["Medidores Reprobados:", str(self.total_failed)],
                ["Tasa de Éxito:", f"{self.success_rate:.1f}%"],
            ]
            
            stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 20))
            
            #  DETALLES POR GRUPO
            for test_group in self.completed_tests:
                story.append(Paragraph(f"GRUPO: {test_group['test_name']}", styles['Heading3']))
                
                # Crear tabla de resultados
                headers = ["#", "Serial", "L. Inicial", "L. Final", "Diferencia", "Vol. Patrón", "Error %", "Estado"]
                data = [headers]
                
                for idx, result in enumerate(test_group["results"], 1):
                    serial_number = result.get("serial_number", result.get("serial", "N/A"))
                    row = [
                        str(idx),
                        serial_number,
                        f"{result.get('initial_reading', 0):.3f}",
                        f"{result.get('final_reading', 0):.3f}",
                        f"{result.get('volume_difference', 0):.3f}",
                        f"{result.get('pattern_volume', 0):.3f}",
                        f"{result.get('error_percentage', 0):.2f}%",
                        "PASA" if result.get("is_passed", False) else "NO PASA"
                    ]
                    data.append(row)
                
                table = Table(data, colWidths=[0.4*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    # Colorear filas según estado
                    *[('BACKGROUND', (0, i+1), (-1, i+1), colors.lightgreen if row[-1] == "PASA" else colors.lightcoral)
                      for i, row in enumerate(data[1:])]
                ]))
                
                story.append(table)
                story.append(Spacer(1, 15))
            
            #  GENERAR PDF
            doc.build(story)
            print(f"[RESULTS_SUMMARY]  PDF generado en: {file_path}")
            
        except ImportError:
            raise Exception("Se requiere instalar 'reportlab' para exportar a PDF:\npip install reportlab")
        except Exception as e:
            raise Exception(f"Error generando PDF: {str(e)}")

    def _start_new_session(self, e):
        """ INICIA NUEVA SESIÓN"""
        print("[RESULTS_SUMMARY]  Iniciando nueva sesión...")
        try:
            from views.batch_registration_view import get_batch_registration_view
            
            self.page.controls.clear()
            new_view = get_batch_registration_view(self.page)
            self.page.controls.append(new_view)
            self.page.update()
            
        except Exception as error:
            print(f"[RESULTS_SUMMARY]  Error iniciando nueva sesión: {error}")
            self._show_notification(" Error iniciando nueva sesión")

    def _go_to_home(self, e):
        """ VA AL INICIO"""
        print("[RESULTS_SUMMARY] 🏠 Volviendo al inicio...")
        try:
            from main import main_menu
            
            self.page.controls.clear()
            main_view = main_menu(self.page)
            self.page.controls.append(main_view)
            self.page.update()
            
        except Exception as error:
            print(f"[RESULTS_SUMMARY]  Error volviendo al inicio: {error}")
            self._show_notification(" Error volviendo al inicio")

    def _print_report(self, e):
        """ IMPRIME EL REPORTE"""
        print("[RESULTS_SUMMARY] 🖨️ Imprimiendo reporte...")
        # TODO: Implementar impresión
        self._show_notification("🖨️ Función de impresión en desarrollo")

    def _show_notification(self, message):
        """ MUESTRA NOTIFICACIÓN"""
        snack_bar = ft.SnackBar(
            content=ft.Text(message),
            duration=3000,
        )
        self.page.overlay.append(snack_bar)
        snack_bar.open = True
        self.page.update()


def get_results_summary_view(page, summary_data):
    """ FUNCIÓN PRINCIPAL PARA OBTENER LA VISTA DE RESUMEN"""
    try:
        print("[RESULTS_SUMMARY]  Creando vista de resumen de resultados...")
        view = ResultsSummaryView(page, summary_data)
        built_view = view.build()
        print("[RESULTS_SUMMARY]  Vista de resumen creada exitosamente")
        return built_view
    except Exception as e:
        print(f"[RESULTS_SUMMARY]  Error creando vista: {e}")
        return ft.Container(
            content=ft.Column([
                ft.Text("Error creando vista de resumen", size=20, color=ft.Colors.RED, weight="bold"),
                ft.Text(f"Error: {str(e)}", size=14, color=ft.Colors.RED_700),
            ], spacing=20, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=40,
            alignment=ft.alignment.center,
            expand=True,
        )